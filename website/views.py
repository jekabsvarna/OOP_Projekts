from flask import Blueprint, render_template, flash, redirect, url_for, request, jsonify
from flask_login import login_required, current_user
from .models import User, Article, Workshop, Enrollment
from fetchLecturers import get_lecturers, get_total_lecturers_count
from fetchStudents import get_students, get_total_students_count
from fetchArticles import get_articles
from werkzeug.security import generate_password_hash
from . import db
from flask import request
import openpyxl, json
from math import ceil
from datetime import datetime
from sqlalchemy.exc import SQLAlchemyError



views = Blueprint('views', __name__)

@views.route('/home_admin')
@login_required
def home_admin():
    db_path = 'instance/database.db'  
    articles = Article.query.all()  
    return render_template("home.html", user=current_user, articles=articles)

@views.route('/admin_lecturers')  
@login_required
def admin_lecturers():
    if current_user.role == 'admin':
        db_path = 'instance/database.db'
        num_lecturers_per_page = request.args.get('num_lecturers', 10, type=int)
        page = request.args.get('page', 1, type=int)
        total_lecturers = get_total_lecturers_count(db_path)  
        total_pages = ceil(total_lecturers / num_lecturers_per_page)
        offset = (page - 1) * num_lecturers_per_page
        lecturers = get_lecturers(db_path, limit=num_lecturers_per_page, offset=offset)
        lecturers_list = [{'id': l[0], 'email': l[1], 'first_name': l[2], 'role': l[3]} for l in lecturers]

        prev_url = url_for('views.admin_lecturers', num_lecturers=num_lecturers_per_page, page=page-1) if page > 1 else None
        next_url = url_for('views.admin_lecturers', num_lecturers=num_lecturers_per_page, page=page+1) if page < total_pages else None

        return render_template("admin_lecturers.html", user=current_user, lecturers=lecturers_list, prev_url=prev_url, next_url=next_url)
    else:
        flash("You are not authorized to access this page.", category="error")
        return redirect(url_for("views.home"))   
    
   
@views.route('/home_lecturer')
@login_required
def home_lecturer():
    if current_user.email.endswith("@rtu.lv") and current_user.email != "admin@rtu.lv":  
        db_path = 'instance/database.db'  
        students = get_students(db_path)
        students_list = [{'id': id, 'email': email, 'first_name': first_name, 'role': role} 
                          for id, email, first_name, role in students]
        return render_template("home_lecturer.html", user=current_user, students=students_list)
    else:
        flash("You are not authorized to access this page.", category="error")
        return redirect(url_for("views.home"))  

@views.route('/home_lecturer_students')
@login_required
def home_lecturer_students():
    if current_user.role == 'lecturer':
        db_path = 'instance/database.db'

        num_students_per_page = request.args.get('num_students', 10, type=int)
        page = request.args.get('page', 1, type=int)

        total_students = get_total_students_count(db_path)
        total_pages = ceil(total_students / num_students_per_page)
        offset = (page - 1) * num_students_per_page

        students = get_students(db_path, limit=num_students_per_page, offset=offset)  
        students_list = [{'id': student[0], 'email': student[1], 'first_name': student[2], 'role': student[3]} for student in students]

        prev_url = url_for('views.home_lecturer_students', num_students=num_students_per_page, page=page-1) if page > 1 else None
        next_url = url_for('views.home_lecturer_students', num_students=num_students_per_page, page=page+1) if page < total_pages else None

        return render_template("home_lecturer_students.html", user=current_user, students=students_list, prev_url=prev_url, next_url=next_url)
    else:
        flash("You are not authorized to access this page.", category="error")
        return redirect(url_for("views.home"))



@views.route('/home_lecturer_articles')
@login_required
def home_lecturer_articles():
    db_path = 'instance/database.db'

    num_articles_per_page = request.args.get('num_articles', 10, type=int)  
    page = request.args.get('page', 1, type=int)  

    total_articles = Article.query.count()  
    total_pages = ceil(total_articles / num_articles_per_page)
    offset = (page - 1) * num_articles_per_page

    articles = Article.query.offset(offset).limit(num_articles_per_page).all()  

    prev_url = url_for('views.home_lecturer_articles', num_articles=num_articles_per_page, page=page-1) if page > 1 else None
    next_url = url_for('views.home_lecturer_articles', num_articles=num_articles_per_page, page=page+1) if page < total_pages else None

    return render_template("home_lecturer_articles.html", user=current_user, articles=articles, prev_url=prev_url, next_url=next_url)


@views.route('/')
@login_required
def home():
    db_path = 'instance/database.db'  
    articles = Article.query.all()  
    return render_template("home.html", user=current_user, articles=articles)

@views.route('/home_user_articles')
@login_required
def home_user_articles():
    num_articles_per_page = request.args.get('num_articles', 10, type=int)  
    page = request.args.get('page', 1, type=int)  

    total_articles = Article.query.count()  
    total_pages = ceil(total_articles / num_articles_per_page)
    offset = (page - 1) * num_articles_per_page

    articles = Article.query.offset(offset).limit(num_articles_per_page).all()  

    prev_url = url_for('views.home_user_articles', num_articles=num_articles_per_page, page=page-1) if page > 1 else None
    next_url = url_for('views.home_user_articles', num_articles=num_articles_per_page, page=page+1) if page < total_pages else None

    return render_template("home_user_articles.html", user=current_user, articles=articles, prev_url=prev_url, next_url=next_url)

@views.route('/add_lecturer', methods=['GET', 'POST'])
@login_required
def add_lecturer():
    if current_user.role == 'admin':
        if request.method == 'POST':
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            email = request.form.get('email')
            password = request.form.get('password')
            
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            
            new_lecturer = User(email=email, first_name=first_name + " " + last_name, 
                                password=password, role='lecturer')
            
            db.session.add(new_lecturer)
            try:
                db.session.commit()
                flash('Lecturer added successfully!', category='success')
            except:
                db.session.rollback()  
                flash('Error adding lecturer.', category='error')
            
            return redirect(url_for('views.home_admin'))

        return render_template('add_lecturer.html', user=current_user)  
    else:
        flash("You are not authorized to perform this action.", category='error')
        return redirect(url_for('views.home'))

@views.route('/import_students', methods=['POST'])
@login_required
def import_students():
    if 'students_xlsx' not in request.files:
        flash('No file part', category='error')
        return redirect(request.url)
    
    file = request.files['students_xlsx']
    
    if file.filename == '':
        flash('No selected file', category='error')
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        try:
            students_data = read_students_xlsx(file)
            save_students_to_database(students_data)
            flash('Students imported successfully!', category='success')
        except Exception as e:
            flash(f'Error importing students: {str(e)}', category='error')
    else:
        flash('Invalid file format', category='error')
    
    return redirect(url_for('views.home_lecturer')) 


def read_students_xlsx(file):
    students_data = []
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 3:  
                email = row[2] if row[2] else ''
                first_name = row[0] if row[0] else ''
                last_name = row[1] if row[1] else ''
                students_data.append({'email': email, 'first_name': first_name, 'last_name': last_name, 'role': 'student'})
            else:
                raise ValueError('Insufficient values in the row')
        return students_data
    except Exception as e:
        raise e

def save_students_to_database(students_data):
    try:
        for student in students_data:
            existing_student = User.query.filter_by(email=student['email']).first()
            if existing_student:
                flash(f'Student with email {student["email"]} already exists.', category='error')
                continue

            full_name = f"{student['first_name']} {student['last_name']}"
            password1 = "securepassword"
            new_student = User(email=student['email'], 
                               first_name=full_name, 
                               password=password1,
                               role=student['role'])
            db.session.add(new_student)
        db.session.commit()
        flash('Students imported successfully!', category='success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing students: {str(e)}', category='error')
        raise e

@views.route('/upload_articles', methods=['POST'])
def upload_articles():
    if 'articles_xlsx' not in request.files:
        flash('No file part', category='error')
        return redirect(request.url)
    
    file = request.files['articles_xlsx']
    
    if file.filename == '':
        flash('No selected file', category='error')
        return redirect(request.url)
    if file and allowed_file(file.filename):
        try:
            article_names = read_xlsx_file(file) 
            save_article_names_to_database(article_names)
            flash('Article names uploaded successfully!', category='success')
        except Exception as e:
            flash(f'Failed to read article names from file: {str(e)}', category='error')
    else:
        flash('Invalid file format', category='error')
    
    return redirect(url_for('views.home_lecturer'))


def read_xlsx_file(file):
    article_names = []
    try:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            article_names.append(row[0])
        return article_names
    except Exception as e:
        print(f"Error reading XLSX file: {e}")
        return None

def save_article_names_to_database(article_names):
    try:
        for name in article_names:
            existing_article = Article.query.filter_by(name=name).first()
            if existing_article:
                flash(f'Article with name "{name}" already exists.', category='error')
                continue

            new_article = Article(name=name)
            db.session.add(new_article)
        db.session.commit()
        flash('Article names uploaded successfully!', category='success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving article names to database: {e}', category='error')
        raise e

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'xlsx'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@views.route('/untake_article', methods=['POST'])
@login_required
def untake_article():
    article_id = request.form.get('article_id')

    article = Article.query.get(article_id)

    if not article:
        return jsonify({'success': False, 'message': 'Article not found.'}), 404

    if article.owner_id != current_user.id:
        return jsonify({'success': False, 'message': 'You are not authorized to untake this article.'}), 403

    article.taken = False
    article.owner_id = None
    db.session.commit()

    return jsonify({'success': True, 'message': 'Article untaken successfully.'}), 200

@views.route('/profile')
@login_required
def profile():
    return render_template("profile.html", user=current_user)

@views.route('/update_name', methods=['POST'])
@login_required
def update_name():
    if request.method == 'POST':
        new_name = request.form.get('name')
        current_user.first_name = new_name
        db.session.commit()
        flash('Name updated successfully!', category='success')
    return redirect(url_for('views.profile'))

@views.route('/update_email', methods=['POST'])
@login_required
def update_email():
    if request.method == 'POST':
        new_email = request.form.get('email')
        current_user.email = new_email
        db.session.commit()
        flash('Email updated successfully!', category='success')
    return redirect(url_for('views.profile'))

@views.route('/update_password', methods=['POST'])
@login_required
def update_password():
    if request.method == 'POST':
        new_password = request.form.get('password')
        current_user.password = new_password
        db.session.commit()
        flash('Password updated successfully!', category='success')
    return redirect(url_for('views.profile'))

@views.route('/update_email', methods=['POST'])
@login_required
def enroll_student(workshop_id):
    student_id = request.form['student_id']
    enrollment = Enrollment(student_id=student_id, workshop_id=workshop_id)
    db.session.add(enrollment)
    db.session.commit()
    return redirect(url_for('workshop_details', workshop_id=workshop_id))

@views.route('/add_workshop', methods=['GET', 'POST'])
@login_required
def add_workshop():
    if request.method == 'POST':
        title = request.form.get('title')
        date_str = request.form.get('date')
        students_json = request.form.get('students', '[]')
        group_data_json = request.form.get('group_data', '[]')

        if not title or not date_str:
            flash('Both title and date must be provided.', 'error')
            students = User.query.filter_by(role='student').all()
            workshops = Workshop.query.all()
            return render_template('add_workshop.html', students=students, workshops=workshops)

        try:
            formatted_date = datetime.strptime(date_str, '%Y-%m-%dT%H:%M')
            selected_students = json.loads(students_json)
            group_data = json.loads(group_data_json) if group_data_json else []

            new_workshop = Workshop(title=title, date=formatted_date, lecturer_id=current_user.id)
            db.session.add(new_workshop)
            db.session.commit()

            for group_index, group_students in enumerate(group_data):
                for student_id in group_students:
                    workshop_id_with_group = f"{new_workshop.id}.{group_index + 1}"
                    enrollment = Enrollment(student_id=student_id, workshop_id=workshop_id_with_group)
                    db.session.add(enrollment)
            db.session.commit()

            flash('Workshop created successfully.', 'success')
            return redirect(url_for('views.home_lecturer'))

        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {e}', 'error')

    workshops = Workshop.query.all()
    students = User.query.filter_by(role='student').all()
    return render_template('add_workshop.html', workshops=workshops, students=students)


@views.route('/delete_student/<int:student_id>', methods=['DELETE'])
@login_required
def delete_student(student_id):
    student = User.query.get_or_404(student_id)
    try:
        db.session.delete(student)
        db.session.commit()
        return jsonify({'message': 'Student deleted successfully'}), 200
    except Exception as e:
        print(f"An error occurred while deleting student: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'An error occurred while deleting student.'}), 500

@views.route('/delete_all_students', methods=['DELETE'])
@login_required
def delete_all_students():
    try:
        students = User.query.filter_by(role='student').all()
        
        for student in students:
            db.session.delete(student)
        
        db.session.commit()
        
        return jsonify({'message': 'All students deleted successfully'}), 200
    except Exception as e:
        print(f"An error occurred while deleting all students: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'An error occurred while deleting all students.'}), 500

@views.route('/delete_all_articles', methods=['DELETE'])
@login_required
def delete_all_articles():
    try:
        Article.query.delete()
        db.session.commit()
        return jsonify({'message': 'All articles deleted successfully'}), 200
    except Exception as e:
        print(f"An error occurred while deleting all articles: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'An error occurred while deleting all articles.'}), 500

@views.route('/delete_article/<int:article_id>', methods=['DELETE'])
@login_required
def delete_article(article_id):
    article = Article.query.get_or_404(article_id)
    try:
        db.session.delete(article)
        db.session.commit()
        return jsonify({'message': 'Article deleted successfully'}), 200
    except Exception as e:
        print(f"An error occurred while deleting the article: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'An error occurred while deleting the article.'}), 500

@views.route('/add_student', methods=['GET', 'POST']) 
@login_required
def add_student():
    if current_user.role == 'lecturer':
        if request.method == 'POST':
            first_name = request.form.get('first_name')
            last_name = request.form.get('last_name')
            email = request.form.get('email')
            password = request.form.get('password')
            
            hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
            
            new_student = User(email=email, first_name=first_name + " " + last_name, 
                                password=password, role='student')
            
            db.session.add(new_student)
            try:
                db.session.commit()
                flash('Student added successfully!', category='success')
            except:
                db.session.rollback()  
                flash('Error adding student.', category='error')
            
            return redirect(url_for('views.home_lecturer_students')) 

        return render_template('add_student.html', user=current_user) 
    else:
        flash("You are not authorized to perform this action.", category='error')
        return redirect(url_for('views.home'))
    
@views.route('/add_article', methods=['GET', 'POST'])
@login_required
def add_article():
    if current_user.role in ['lecturer', 'admin']:  
        if request.method == 'POST':
            article_name = request.form.get('article_name')
            new_article = Article(name=article_name)  
            db.session.add(new_article)
            try:
                db.session.commit()
                flash('Article added successfully!', category='success')
                return redirect(url_for('views.home_lecturer_articles'))
            except Exception as e:
                db.session.rollback()
                flash(f'Error adding article: {str(e)}', category='error')
                return render_template('add_article.html')
        return render_template('add_article.html')
    else:
        flash("You are not authorized to perform this action.", category='error')
        return redirect(url_for('views.home'))

